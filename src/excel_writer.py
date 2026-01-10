"""Excel output with multiple tabs for different report types."""

from pathlib import Path
from typing import Dict, List, Any, Optional
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Create multi-tab Excel files from parsed financial data."""

    def __init__(self, output_path: Path):
        """
        Initialize Excel writer.

        Args:
            output_path: Path for output .xlsx file
        """
        self.output_path = Path(output_path)
        self.output_path.parent.mkdir(parents=True, exist_ok=True)

        # Import here to allow graceful failure
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            self.openpyxl = openpyxl
            self.styles = {
                'Font': Font,
                'PatternFill': PatternFill,
                'Alignment': Alignment,
                'Border': Border,
                'Side': Side
            }
        except ImportError:
            raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

        self.workbook = self.openpyxl.Workbook()
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        self.sheets = {}

    def _get_or_create_sheet(self, name: str):
        """Get existing sheet or create new one."""
        # Excel sheet names max 31 chars
        name = name[:31]
        if name not in self.sheets:
            self.sheets[name] = self.workbook.create_sheet(title=name)
        return self.sheets[name]

    def _apply_header_style(self, cell):
        """Apply header styling to a cell."""
        Font = self.styles['Font']
        PatternFill = self.styles['PatternFill']
        Alignment = self.styles['Alignment']

        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    def _apply_currency_format(self, cell):
        """Apply currency format to a cell."""
        cell.number_format = '$#,##0.00'

    def _auto_adjust_columns(self, sheet):
        """Auto-adjust column widths based on content."""
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils import get_column_letter

        for col_idx, column_cells in enumerate(sheet.columns, 1):
            max_length = 0
            for cell in column_cells:
                # Skip merged cells
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            # Cap at reasonable width
            adjusted_width = min(max_length + 2, 50) if max_length > 0 else 10
            column_letter = get_column_letter(col_idx)
            sheet.column_dimensions[column_letter].width = adjusted_width

    def add_balance_sheet(self, data: List[Dict[str, Any]], sheet_name: str = "Balance Sheet"):
        """
        Add balance sheet data to workbook.

        Expected data format:
        [
            {
                "account_code": "1001",
                "account_name": "PPB Builder Bond",
                "category": "Assets",
                "subcategory": "Operating Funds",
                "current_balance": 21398.80,
                "prior_balance": 21394.40,
                "change": 4.40
            },
            ...
        ]
        """
        sheet = self._get_or_create_sheet(sheet_name)

        headers = [
            "Account Code", "Account Name", "Category", "Subcategory",
            "Current Balance", "Prior Balance", "Change"
        ]

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        # Write data
        for row_idx, record in enumerate(data, 2):
            sheet.cell(row=row_idx, column=1, value=record.get('account_code', ''))
            sheet.cell(row=row_idx, column=2, value=record.get('account_name', ''))
            sheet.cell(row=row_idx, column=3, value=record.get('category', ''))
            sheet.cell(row=row_idx, column=4, value=record.get('subcategory', ''))

            for col, field in [(5, 'current_balance'), (6, 'prior_balance'), (7, 'change')]:
                cell = sheet.cell(row=row_idx, column=col, value=record.get(field, 0))
                self._apply_currency_format(cell)

        self._auto_adjust_columns(sheet)
        logger.info(f"Added {len(data)} rows to {sheet_name}")

    def add_disbursements(self, data: List[Dict[str, Any]], sheet_name: str = "Disbursements"):
        """
        Add check disbursement data to workbook.

        Expected data format:
        [
            {
                "check_number": "00200284",
                "check_date": "2025-11-03",
                "vendor": "Associa Hill Country",
                "account_code": "7040",
                "account_name": "Management Fees",
                "description": "Management Fee",
                "amount": 805.00,
                "category": "Management"
            },
            ...
        ]
        """
        sheet = self._get_or_create_sheet(sheet_name)

        headers = [
            "Check #", "Date", "Vendor", "Account Code", "Account Name",
            "Description", "Amount", "Category"
        ]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        for row_idx, record in enumerate(data, 2):
            sheet.cell(row=row_idx, column=1, value=record.get('check_number', ''))
            sheet.cell(row=row_idx, column=2, value=record.get('check_date', ''))
            sheet.cell(row=row_idx, column=3, value=record.get('vendor', ''))
            sheet.cell(row=row_idx, column=4, value=record.get('account_code', ''))
            sheet.cell(row=row_idx, column=5, value=record.get('account_name', ''))
            sheet.cell(row=row_idx, column=6, value=record.get('description', ''))

            amount_cell = sheet.cell(row=row_idx, column=7, value=record.get('amount', 0))
            self._apply_currency_format(amount_cell)

            sheet.cell(row=row_idx, column=8, value=record.get('category', ''))

        self._auto_adjust_columns(sheet)
        logger.info(f"Added {len(data)} rows to {sheet_name}")

    def add_invoices(self, data: List[Dict[str, Any]], sheet_name: str = "Invoices"):
        """
        Add invoice data to workbook.

        Expected data format:
        [
            {
                "invoice_id": "890931",
                "invoice_date": "2025-10-24",
                "vendor": "Associa OnCall",
                "description": "Gate issue/phone line",
                "amount": 428.68,
                "source_page": 61,
                "ocr_confidence": "high"
            },
            ...
        ]
        """
        sheet = self._get_or_create_sheet(sheet_name)

        headers = [
            "Invoice ID", "Date", "Vendor", "Description",
            "Amount", "Source Page", "OCR Confidence"
        ]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        for row_idx, record in enumerate(data, 2):
            sheet.cell(row=row_idx, column=1, value=record.get('invoice_id', ''))
            sheet.cell(row=row_idx, column=2, value=record.get('invoice_date', ''))
            sheet.cell(row=row_idx, column=3, value=record.get('vendor', ''))
            sheet.cell(row=row_idx, column=4, value=record.get('description', ''))

            amount_cell = sheet.cell(row=row_idx, column=5, value=record.get('amount', 0))
            self._apply_currency_format(amount_cell)

            sheet.cell(row=row_idx, column=6, value=record.get('source_page', ''))
            sheet.cell(row=row_idx, column=7, value=record.get('ocr_confidence', ''))

        self._auto_adjust_columns(sheet)
        logger.info(f"Added {len(data)} rows to {sheet_name}")

    def add_investments(self, data: List[Dict[str, Any]], sheet_name: str = "Investments"):
        """
        Add investment/bank account data to workbook.

        Expected data format:
        [
            {
                "account_code": "1001",
                "account_name": "PPB Builder Bond",
                "institution": "Pacific Premier Bank",
                "account_number": "****3118",
                "type": "Money Market",
                "balance": 21398.80,
                "rate": 0.25
            },
            ...
        ]
        """
        sheet = self._get_or_create_sheet(sheet_name)

        headers = [
            "Account Code", "Account Name", "Institution",
            "Account #", "Type", "Balance", "Rate %"
        ]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        for row_idx, record in enumerate(data, 2):
            sheet.cell(row=row_idx, column=1, value=record.get('account_code', ''))
            sheet.cell(row=row_idx, column=2, value=record.get('account_name', ''))
            sheet.cell(row=row_idx, column=3, value=record.get('institution', ''))
            sheet.cell(row=row_idx, column=4, value=record.get('account_number', ''))
            sheet.cell(row=row_idx, column=5, value=record.get('type', ''))

            balance_cell = sheet.cell(row=row_idx, column=6, value=record.get('balance', 0))
            self._apply_currency_format(balance_cell)

            rate_cell = sheet.cell(row=row_idx, column=7, value=record.get('rate', 0))
            rate_cell.number_format = '0.00%'

        self._auto_adjust_columns(sheet)
        logger.info(f"Added {len(data)} rows to {sheet_name}")

    def add_bank_reconciliation(self, data: List[Dict[str, Any]], sheet_name: str = "Bank Reconciliation"):
        """
        Add bank reconciliation data to workbook.

        Expected data format:
        [
            {
                "account_code": "1011",
                "account_name": "HAR OPER #1137",
                "account_type": "Operating",
                "balance_per_bank": 3522.75,
                "total_outstanding_deposits": 12300.00,
                "total_outstanding_checks": -410.50,
                "ending_balance_gl": 15412.25,
                "difference": 0.00,
                "is_reconciled": true
            },
            ...
        ]
        """
        sheet = self._get_or_create_sheet(sheet_name)

        headers = [
            "Account Code", "Account Name", "Type", "Bank Balance",
            "Outstanding Deposits", "Outstanding Checks", "GL Balance",
            "Difference", "Reconciled"
        ]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        for row_idx, record in enumerate(data, 2):
            sheet.cell(row=row_idx, column=1, value=record.get('account_code', ''))
            sheet.cell(row=row_idx, column=2, value=record.get('account_name', ''))
            sheet.cell(row=row_idx, column=3, value=record.get('account_type', ''))

            for col, field in [
                (4, 'balance_per_bank'),
                (5, 'total_outstanding_deposits'),
                (6, 'total_outstanding_checks'),
                (7, 'ending_balance_gl'),
                (8, 'difference')
            ]:
                cell = sheet.cell(row=row_idx, column=col, value=record.get(field, 0))
                self._apply_currency_format(cell)

            reconciled = record.get('is_reconciled', False)
            sheet.cell(row=row_idx, column=9, value="Yes" if reconciled else "No")

        self._auto_adjust_columns(sheet)
        logger.info(f"Added {len(data)} rows to {sheet_name}")

    def add_accounts_receivable(self, data: List[Dict[str, Any]], sheet_name: str = "Accounts Receivable"):
        """
        Add accounts receivable / delinquency data to workbook.

        Expected data format:
        [
            {
                "account_id": "00339-2087",
                "name": "Brad S. Rose",
                "address": "1202 Brads Flight",
                "section": "delinquent",
                "day_30": 71.01,
                "day_31_60": 235.80,
                "day_61_90": 40.59,
                "day_91_120": 40.38,
                "day_120_plus": 4069.35,
                "total_balance": 4457.13
            },
            ...
        ]
        """
        sheet = self._get_or_create_sheet(sheet_name)

        headers = [
            "Account ID", "Name", "Address", "Status",
            "Current", "31-60 Days", "61-90 Days", "91-120 Days",
            "120+ Days", "Total Balance"
        ]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        for row_idx, record in enumerate(data, 2):
            sheet.cell(row=row_idx, column=1, value=record.get('account_id', ''))
            sheet.cell(row=row_idx, column=2, value=record.get('name', ''))
            sheet.cell(row=row_idx, column=3, value=record.get('address', ''))
            sheet.cell(row=row_idx, column=4, value=record.get('section', '').title())

            for col, field in [
                (5, 'day_30'),
                (6, 'day_31_60'),
                (7, 'day_61_90'),
                (8, 'day_91_120'),
                (9, 'day_120_plus'),
                (10, 'total_balance')
            ]:
                cell = sheet.cell(row=row_idx, column=col, value=record.get(field, 0))
                self._apply_currency_format(cell)

        self._auto_adjust_columns(sheet)
        logger.info(f"Added {len(data)} rows to {sheet_name}")

    def add_income_statement(self, data: List[Dict[str, Any]], sheet_name: str = "Income Statement"):
        """
        Add income statement data to workbook.

        Expected data format:
        [
            {
                "account_code": "4000",
                "account_name": "Residential Assessments",
                "section": "Income",
                "category": "Assessment Income",
                "is_total": false,
                "current_actual": 0.00,
                "current_budget": 0.00,
                "ytd_actual": 123516.80,
                "ytd_budget": 124920.00,
                "ytd_variance": -1403.20,
                "annual_budget": 124920.00,
                "budget_remaining": 1403.20
            },
            ...
        ]
        """
        sheet = self._get_or_create_sheet(sheet_name)

        headers = [
            "Account", "Name", "Section", "Category",
            "Curr Actual", "Curr Budget", "Curr Var",
            "YTD Actual", "YTD Budget", "YTD Var",
            "Annual Budget", "Remaining"
        ]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        for row_idx, record in enumerate(data, 2):
            sheet.cell(row=row_idx, column=1, value=record.get('account_code', ''))
            sheet.cell(row=row_idx, column=2, value=record.get('account_name', ''))
            sheet.cell(row=row_idx, column=3, value=record.get('section', ''))
            sheet.cell(row=row_idx, column=4, value=record.get('category', ''))

            # Highlight total rows
            is_total = record.get('is_total', False)

            for col, field in [
                (5, 'current_actual'),
                (6, 'current_budget'),
                (7, 'current_variance'),
                (8, 'ytd_actual'),
                (9, 'ytd_budget'),
                (10, 'ytd_variance'),
                (11, 'annual_budget'),
                (12, 'budget_remaining')
            ]:
                cell = sheet.cell(row=row_idx, column=col, value=record.get(field, 0))
                self._apply_currency_format(cell)
                if is_total:
                    cell.font = self.styles['Font'](bold=True)

        self._auto_adjust_columns(sheet)
        logger.info(f"Added {len(data)} rows to {sheet_name}")

    def add_expense_trend(self, data: List[Dict[str, Any]], sheet_name: str = "Expense Trend"):
        """
        Add expense trend data (monthly breakdown) to workbook.

        Expected data format:
        [
            {
                "account_code": "5000",
                "account_name": "Administrative Supplies",
                "category": "Administrative",
                "is_total": false,
                "jan": 20.00,
                "feb": 0.00,
                "mar": 72.00,
                ... (all months)
                "full_year_actual": 1076.00,
                "total_budget": 1100.00
            },
            ...
        ]
        """
        sheet = self._get_or_create_sheet(sheet_name)

        headers = [
            "Account", "Name", "Category",
            "Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov",
            "Full Year", "Budget", "Variance"
        ]

        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)

        for row_idx, record in enumerate(data, 2):
            sheet.cell(row=row_idx, column=1, value=record.get('account_code', ''))
            sheet.cell(row=row_idx, column=2, value=record.get('account_name', ''))
            sheet.cell(row=row_idx, column=3, value=record.get('category', ''))

            # Highlight total rows
            is_total = record.get('is_total', False)

            # Monthly columns
            month_fields = ['jan', 'feb', 'mar', 'apr', 'may', 'jun',
                           'jul', 'aug', 'sep', 'oct', 'nov']
            for col_offset, month in enumerate(month_fields):
                cell = sheet.cell(row=row_idx, column=4 + col_offset,
                                 value=record.get(month, 0))
                self._apply_currency_format(cell)
                if is_total:
                    cell.font = self.styles['Font'](bold=True)

            # Full Year Actual (column 15)
            full_year = record.get('full_year_actual', 0)
            cell = sheet.cell(row=row_idx, column=15, value=full_year)
            self._apply_currency_format(cell)
            if is_total:
                cell.font = self.styles['Font'](bold=True)

            # Budget (column 16)
            budget = record.get('total_budget', 0)
            cell = sheet.cell(row=row_idx, column=16, value=budget)
            self._apply_currency_format(cell)
            if is_total:
                cell.font = self.styles['Font'](bold=True)

            # Variance (column 17) = Budget - Actual
            variance = budget - full_year if budget and full_year else 0
            cell = sheet.cell(row=row_idx, column=17, value=variance)
            self._apply_currency_format(cell)
            if is_total:
                cell.font = self.styles['Font'](bold=True)

        self._auto_adjust_columns(sheet)
        logger.info(f"Added {len(data)} rows to {sheet_name}")

    def add_summary(self, summary_data: Dict[str, Any], sheet_name: str = "Summary"):
        """
        Add a summary sheet with key metrics.

        Expected data format:
        {
            "report_date": "November 30, 2025",
            "total_assets": 124033.38,
            "total_liabilities": 23895.46,
            "net_equity": 100137.92,
            "operating_funds": 14458.25,
            "reserve_funds": 99143.16,
            "accounts_receivable": 8594.97,
            "monthly_expenses": 15000.00,
            "checks_written": 25
        }
        """
        sheet = self._get_or_create_sheet(sheet_name)

        # Title
        title_cell = sheet.cell(row=1, column=1, value="Financial Summary")
        title_cell.font = self.styles['Font'](bold=True, size=14)
        sheet.merge_cells('A1:B1')

        # Report date
        sheet.cell(row=2, column=1, value="Report Date:")
        sheet.cell(row=2, column=2, value=summary_data.get('report_date', ''))

        # Generated timestamp
        sheet.cell(row=3, column=1, value="Generated:")
        sheet.cell(row=3, column=2, value=datetime.now().strftime('%Y-%m-%d %H:%M'))

        # Key metrics
        metrics = [
            ("Total Assets", 'total_assets'),
            ("Total Liabilities", 'total_liabilities'),
            ("Net Equity", 'net_equity'),
            ("Operating Funds", 'operating_funds'),
            ("Reserve Funds", 'reserve_funds'),
            ("Accounts Receivable", 'accounts_receivable'),
            ("Monthly Expenses", 'monthly_expenses'),
        ]

        row = 5
        for label, key in metrics:
            sheet.cell(row=row, column=1, value=label)
            cell = sheet.cell(row=row, column=2, value=summary_data.get(key, 0))
            self._apply_currency_format(cell)
            row += 1

        # Count metrics
        sheet.cell(row=row, column=1, value="Checks Written")
        sheet.cell(row=row, column=2, value=summary_data.get('checks_written', 0))

        self._auto_adjust_columns(sheet)
        logger.info(f"Added summary sheet")

    def add_raw_data(self, data: List[List[Any]], sheet_name: str, headers: Optional[List[str]] = None):
        """
        Add arbitrary tabular data to workbook.

        Args:
            data: List of rows (each row is a list of values)
            sheet_name: Name for the sheet
            headers: Optional header row
        """
        sheet = self._get_or_create_sheet(sheet_name)

        start_row = 1
        if headers:
            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                self._apply_header_style(cell)
            start_row = 2

        for row_idx, row_data in enumerate(data, start_row):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        self._auto_adjust_columns(sheet)
        logger.info(f"Added {len(data)} rows to {sheet_name}")

    def save(self):
        """Save the workbook to disk."""
        self.workbook.save(self.output_path)
        logger.info(f"Saved workbook to {self.output_path}")

    def close(self):
        """Close the workbook."""
        self.workbook.close()
